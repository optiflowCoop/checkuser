#!/usr/bin/env python3
"""
Populate LOCATION_SITE in workbook from persongroupview files (output/raw/*_persongroupview.txt).
"""
import csv
from pathlib import Path
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent
RAW_DIR = ROOT / 'output' / 'raw'
WORKBOOK = ROOT / 'output' / 'reports' / 'maximo_risk_and_optimization_workbook.xlsx'

def parse_pgv_file(path, mapping):
    with open(path, encoding='utf-8', errors='replace') as f:
        for line in f:
            line = line.rstrip('\n')
            if not line:
                continue
            # Skip header/meta lines
            if line.startswith('IBM') or line.startswith('>') or line.startswith('FetchAll') or line.startswith('  CSV_ROW'):
                continue
            # Now we expect CSV_ROW like: personid,...
            # split on comma (SELECT replaced internal commas)
            cols = line.split(',')
            if not cols:
                continue
            userid = cols[0].strip().upper()
            if not userid:
                continue
            # locationsite is at index 16 (0-based) if present according to SELECT mapping
            loc_site = None
            if len(cols) >= 17:
                cs = cols[16].strip()
                if cs:
                    loc_site = cs
            # fallback: search for a token that looks like N06/N08/ODN etc
            if not loc_site:
                for c in cols[10:30]:
                    if not c:
                        continue
                    v = c.strip().upper()
                    if v.startswith('N0') or v.startswith('ODN') or 'BASE' in v or v in ('ONSHORE','OFFSHORE'):
                        loc_site = c.strip()
                        break
            if loc_site:
                mapping[userid] = loc_site

def build_mapping():
    mapping = {}
    for pgv in RAW_DIR.glob('*_persongroupview.txt'):
        parse_pgv_file(pgv, mapping)
    return mapping

def main():
    mapping = build_mapping()
    print(f'Found {len(mapping)} person -> locationsite mappings from persongroupview')

    if not WORKBOOK.exists():
        print('Workbook not found:', WORKBOOK)
        return

    wb = load_workbook(WORKBOOK)
    ws = wb['2_LicenseDecisionPlan']
    headers = [cell.value for cell in ws[1]]
    try:
        ls_col = headers.index('LOCATION_SITE') + 1
    except ValueError:
        print('LOCATION_SITE header not found in workbook')
        return

    # also load usage CSV to fallback
    usage_csv = ROOT / 'output' / 'consolidated' / 'usage_analysis_phase3.csv'
    usage_map = {}
    if usage_csv.exists():
        with open(usage_csv, encoding='utf-8-sig', errors='replace') as f:
            reader = csv.DictReader(f)
            for row in reader:
                uid = str(row.get('USERID','')).strip().upper()
                if uid:
                    usage_map[uid] = row.get('LOCATION','')

    updated = 0
    for r in range(2, ws.max_row+1):
        uid = (ws.cell(row=r, column=1).value or '').strip().upper()
        if not uid:
            continue
        new_loc = mapping.get(uid)
        if not new_loc:
            new_loc = usage_map.get(uid)
        if new_loc:
            ws.cell(row=r, column=ls_col).value = new_loc
            updated += 1

    print(f'Updating workbook: {updated} rows')
    # save backup
    bak = WORKBOOK.with_name(WORKBOOK.stem + '_backup.xlsx')
    wb.save(bak)
    wb.save(WORKBOOK)
    print('Saved workbook and backup')

if __name__ == '__main__':
    main()
