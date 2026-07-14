# generate_risk_report.py (Orchestrator)
import sys
import csv
import json
from pathlib import Path
from collections import Counter, defaultdict
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.hyperlink import Hyperlink

if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')
    sys.stderr.reconfigure(encoding='utf-8', errors='replace')

# --- CRITICAL FIX: Add the project root to sys.path ---
ROOT = Path(__file__).resolve().parent.parent
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

# Import modularized components
from scripts.config import get_app_points_config, get_entitlement_keywords, get_critical_titles, get_foresea_domains
from scripts.domain.user import build_user_profiles, get_user_domain_category
from scripts.services.analysis import analyze_governance, analyze_title_divergences
from scripts.services.app_points import simulate_app_points
from scripts.reporting.html_builder import build_html_structure
from scripts.reporting.html_helpers import fmt_br, render_table
from scripts.analysis.entitlement import calculate_app_points

# --- NOVA IMPORTAÇÃO CORRIGIDA ---
from scripts.domain.identity_analyzer import get_unique_users_data
# --- NOVA IMPORTAÇÃO: SANITY ANALYZER ---
from scripts.domain.sanity_analyzer import analyze_sanity
from scripts.domain.migration_advisor import analyze_migration
from scripts.domain.allocation_analyzer import analyze_allocation
from scripts.domain.security_audit import analyze_security_audit
from scripts.domain.group_baseline import analyze_group_baseline
from scripts.domain.role_standardization import analyze_role_standardization
from scripts.domain.license_reconciliation import analyze_license_reconciliation


# --- Constants ---
IN_DIR = ROOT / 'output' / 'consolidated'
OUT_DIR = ROOT / 'output' / 'reports'
OUT_DIR.mkdir(parents=True, exist_ok=True)

# --- Data Loading ---
def load_all_data():
    """Loads all necessary CSV files from the consolidated directory."""
    return {
        "identities": load_csv(IN_DIR / 'consolidated_user_identity.csv'),
        "access_rows": load_csv(IN_DIR / 'consolidated_user_access_normalized.csv') or load_csv(IN_DIR / 'consolidated_user_access.csv'),
        "cross_env": load_csv(IN_DIR / 'cross_env_userid_reuse.csv'),
        "login_conflicts": load_csv(IN_DIR / 'login_conflicts.csv'),
        "worklist": load_csv(IN_DIR / 'identity_collisions_enriched.csv'),
        "emails": load_csv(IN_DIR / 'consolidated_email.csv'),
        "persons": load_csv(IN_DIR / 'consolidated_person.csv') + load_person_supplements(),
        "persongroupview": load_csv(IN_DIR / 'consolidated_persongroupview.csv'),
        "groupuser": load_csv(IN_DIR / 'consolidated_groupuser.csv'),
        "maxuserstatus": load_csv(IN_DIR / 'consolidated_maxuserstatus.csv'),
        # --- NOVAS FONTES: AD e Maximo ---
        "ad_users": load_csv(IN_DIR / 'consolidated_ad_users.csv'),
        "maximo_users": load_csv(IN_DIR / 'consolidated_maximo_users.csv'),
    }

def detect_delimiter(path: Path):
    """Detecta automaticamente o delimitador de um CSV."""
    with path.open('r', encoding='utf-8-sig') as f:
        first_line = f.readline()
    if ';' in first_line:
        return ';'
    return ','

def load_csv(path: Path):
    """Helper to load a single CSV, returning an empty list if not found."""
    if not path.exists(): return []
    delim = detect_delimiter(path)
    with path.open('r', encoding='utf-8-sig', newline='') as f:
        return list(csv.DictReader(f, delimiter=delim))

def load_person_supplements():
    """Loads supplemental PERSON snapshots kept in the knowledge base folder."""
    rows = []
    base_dir = ROOT / 'Base Conhecimento' / 'Base'
    for path in base_dir.glob('PERSON_*.csv'):
        rows.extend(load_csv(path))
    return rows

def write_license_decision_plan(rows):
    """Writes an auditable CSV with the final license recommendation per user."""
    if not rows:
        return

    fieldnames = [
        'USERID', 'DISPLAYNAME', 'ENTITLEMENT', 'LICENSE_MODEL', 'APP_POINTS',
        'EMAIL', 'DOMAIN_CATEGORY', 'MIGRATION_SCOPE', 'OPERATIONAL_PRESENCE',
        'LOCATION_SITE', 'USAGE_PROFILE', 'OPTIMIZATION_REC',
        'OPTIMIZATION_REASON', 'LOGIN_COUNT_90D', 'LOGIN_COUNT_60D', 'DAYS_SINCE_LAST',
        'FACTOR_P50', 'FACTOR_P95', 'FACTOR_P100', 'TITLES', 'ACTIVE_HOURS'
    ]
    out_path = IN_DIR / 'license_decision_plan.csv'
    with out_path.open('w', encoding='utf-8-sig', newline='') as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction='ignore')
        writer.writeheader()
        for row in rows:
            if 'ACTIVE_HOURS' in row and isinstance(row['ACTIVE_HOURS'], list):
                row['ACTIVE_HOURS'] = '|'.join(row['ACTIVE_HOURS'])
            writer.writerow(row)
    print(f'✓ WROTE {out_path.name}')

def build_master_rows(license_rows, sanity_data, migration_data, allocation_data):
    """Consolida numa única linha por usuário os dados de AD, Maximo, licença,
    saneamento, migração e alocação — a mesma finalidade da antiga "aba 2"
    (todos os dados juntos para validação manual), reconstruída a partir dos
    resultados já calculados pelas outras análises do pipeline.
    """
    sanity_data = sanity_data or {}
    maximo_by_userid = sanity_data.get('maximo_by_userid', {})
    ad_by_email = sanity_data.get('ad_by_email', {})
    ad_disabled_by_email = sanity_data.get('ad_disabled_by_email', {})

    alert_by_userid = {}
    alert_by_email = {}
    for d in sanity_data.get('analises', {}).get('ad_disabled_ativos_maximo', []):
        email = str(d.get('email', '')).strip().lower()
        if email:
            alert_by_email[email] = d
        for uid in str(d.get('maximo_userids', '')).split(' | '):
            uid = uid.strip().upper()
            if uid:
                alert_by_userid[uid] = d

    migration_by_userid = {}
    for r in migration_data or []:
        uid = str(r.get('userid', '')).strip().upper()
        if uid and uid not in migration_by_userid:
            migration_by_userid[uid] = r

    allocation_rows = (allocation_data or {}).get('analises', []) if isinstance(allocation_data, dict) else (allocation_data or [])
    allocation_by_userid = {}
    for a in allocation_rows:
        uid = str(a.get('userid', '')).strip().upper()
        if uid:
            allocation_by_userid[uid] = a

    rows = []
    for lr in license_rows:
        uid = str(lr.get('USERID', '')).strip().upper()
        email = str(lr.get('EMAIL', '')).strip().lower()
        mx = maximo_by_userid.get(uid, {})
        envs_total = sorted(e for e in mx.get('envs', []) if e)
        envs_ativos = sorted(
            e for e, s in mx.get('env_status', {}).items()
            if s.upper() in ('ACTIVE', 'ATIVO', 'ENABLED')
        )

        if email and email in ad_by_email:
            status_ad = 'ATIVO'
        elif email and email in ad_disabled_by_email:
            status_ad = 'DESATIVADO'
        elif email:
            status_ad = 'NAO ENCONTRADO NO AD'
        else:
            status_ad = 'SEM EMAIL (NAO COMPARAVEL)'

        alerta = alert_by_userid.get(uid) or (alert_by_email.get(email) if email else None)
        migr = migration_by_userid.get(uid)
        aloc = allocation_by_userid.get(uid)

        rows.append({
            'USERID': uid,
            'NOME': lr.get('DISPLAYNAME', ''),
            'EMAIL': lr.get('EMAIL', ''),
            'DOMINIO': lr.get('DOMAIN_CATEGORY', ''),
            'TYPE': lr.get('TYPE', ''),
            'GRUPOS_MAXIMO': lr.get('GROUPS', ''),
            'STATUS_AD': status_ad,
            'AMBIENTES_MAXIMO_TOTAL': '; '.join(envs_total),
            'AMBIENTES_MAXIMO_ATIVOS': '; '.join(envs_ativos),
            'AMBIENTES_ATIVOS_DE_TOTAL': f'{len(envs_ativos)}/{len(envs_total)}' if envs_total else '',
            'STATUS_MAXIMO': '; '.join(sorted(mx.get('statuses', []))),
            'ENTITLEMENT': lr.get('ENTITLEMENT', ''),
            'LICENSE_MODEL': lr.get('LICENSE_MODEL', ''),
            'APP_POINTS': lr.get('APP_POINTS', ''),
            'LOCATION_SITE': lr.get('LOCATION_SITE', ''),
            'USAGE_PROFILE': lr.get('USAGE_PROFILE', ''),
            'LOGIN_COUNT_90D': lr.get('LOGIN_COUNT_90D', ''),
            'DAYS_SINCE_LAST': lr.get('DAYS_SINCE_LAST', ''),
            'OPTIMIZATION_REC': lr.get('OPTIMIZATION_REC', ''),
            'OPTIMIZATION_REASON': lr.get('OPTIMIZATION_REASON', ''),
            'ALERTA_AD_DESATIVADO_MAXIMO_ATIVO': (
                f"CRITICO: {alerta.get('qtd_envs_ativos_de_total', '')} ambientes ativos"
                if alerta else ''
            ),
            'MIGRACAO_TIPO': migr.get('tipo', '') if migr else '',
            'MIGRACAO_ACAO': migr.get('acao', '') if migr else '',
            'ALOCACAO_PRINCIPAL': aloc.get('allocation_primary', '') if aloc else '',
            'ALOCACAO_SUGERIDA': '; '.join(aloc.get('suggested_accounts', [])) if aloc else '',
            'TITLES': lr.get('TITLES', ''),
        })
    return rows


def add_index_sheet(wb, add_sheet):
    """Cria a aba de índice/navegação (descrição + contagem de linhas de cada
    aba) e move-a para a primeira posição do workbook."""
    descriptions = {
        '1_VisaoExecutiva': 'Resumo executivo: contagens gerais, distribuicao de licencas e dominios.',
        '2_Master_Analise': 'CONSOLIDADO MESTRE: 1 linha por usuario cruzando AD, Maximo, licenca e alertas de saneamento. Use para validacao manual.',
        '3_LicenseDecisionPlan': 'Plano detalhado de decisao de licenca por usuario (entitlement, modelo, AppPoints, uso).',
        '4_RevisarSemDominio': 'Usuarios sem email/dominio valido - revisar antes de contabilizar licenca.',
        '5_ReusoUSERID_CrossEnv': 'USERIDs reutilizados entre ambientes Maximo diferentes.',
        '6_ConflitosLoginID': 'Conflitos de login (mesmo LOGINID usado por multiplas pessoas/USERIDs).',
        '7_FilaSaneamento': 'Fila de identidades para saneamento manual (hipoteses de colisao).',
        '9_Saneamento_Resumo': 'Resumo estatistico do saneamento de identidades AD x Maximo.',
        '10_Saneamento_AD_Achados': 'TODOS os achados AD x Maximo numa so aba — filtre ACHADO: AD_DESATIVADO_MAS_ATIVO (critico, vermelho), DIVERGENCIA_NOME, MULTIPLOS_USERIDS, MATCH_POR_PREFIXO, AD_SEM_MATCH_MAXIMO, MAXIMO_SEM_EMAIL_COM_AD, DIVERGENCIA_DOMINIO.',
        '11_Migracao_MAS9': 'ENTREGAVEL DE MIGRACAO: por usuario — migrar?, grupos Maximo atuais, grupo recomendado MAS 9, cargo, pares ambiente:status. Resumo por tipo na aba 17.',
        '12_Alocacao_Resumo': 'Resumo do saneamento de alocacao de contas por ambiente (Maximo 9).',
        '13_Alocacao_Sugestao': 'Sugestao detalhada de alocacao de contas por ambiente, por usuario.',
        '14_Maximo_Usuarios_Ativos': 'Lista unica de todos os usuarios ativos no Maximo (todos os ambientes).',
        '15_Acessos_por_Perfil': 'Contagem de usuarios ativos agrupados por tipo de perfil (TYPE).',
        '16_Auditoria_Acesso': 'Data de concessao de acesso (auditoria) por usuario.',
        '17_Resumo_Consolidado': 'RESUMO EXECUTIVO em secoes: Migracao MAS 9 (contagens por tipo), Cenario Conciliado (P95/P100 vs teto 1200) e SoD em Compras (conflitos + evidencias em USD).',
        '18_SoD_Grupos': 'SoD Nivel 1: grupos que sozinhos concedem emissao E aprovacao na mesma app. Filtre a coluna APP para ver so PR, PO ou Requisicao Simplificada.',
        '19_SoD_Pessoas': 'SoD Nivel 2 + governanca: pessoas com o conflito emissor/aprovador E membros de MAXADMIN, na mesma aba (filtre a coluna TIPO). Com cargo e papel recomendado.',
        '20_SoD_Evidencias': 'EVIDENCIAS das 3 camadas numa so aba (filtre CAMADA): SUBMETEU_E_APROVOU (WFTRANSACTION), AUTOAPROVACAO (solicitante real=aprovador) e APROVOU_PR_GEROU_PO (cadeia PR->PO). Valores em USD, por PR unica.',
        '21_Cargo_x_Grupos': 'MODELO DE ACESSO POR CARGO: linhas UNIDADE=TODAS sao o padrao global recomendado MAS 9 (material para terceirizada); linhas por unidade sao o baseline real observado hoje.',
        '22_Perfil_Cargo_Desvios': 'Pessoas cujos grupos desviam do padrao do proprio cargo na mesma unidade (excesso = risco, falta = possivel bloqueio).',
        '23_Grupos_Duplicados': 'MATERIAL PARA TERCEIRIZADA: grupos com permissao >=95% identica sob nomes diferentes — candidatos a fusao. Alertas: privilegio divergente (ADM=leitura) e papeis nominalmente distintos.',
        '24_Cenario_Conciliado_Usuarios': 'DIMENSIONAMENTO MAS 9 por usuario: conciliacao AD (email/prefixo/nome), presenca % em 90d, licenca economica vs final, custo. POPULACAO=TERCEIRO_ATIVO em amarelo. Resumo na aba 17.',
    }
    rows = []
    for name in wb.sheetnames:
        ws = wb[name]
        rows.append({
            'Aba': name,
            'Descricao': descriptions.get(name, ''),
            'Linhas de Dados': max(ws.max_row - 1, 0),
        })

    def index_highlight(row, idx):
        if str(row.get('Aba', '')).startswith('11_'):
            return 'fecaca'  # achado crítico em destaque também no índice
        return 'f8fafc' if idx % 2 else None  # zebra

    add_sheet('0_Indice', ['Aba', 'Descricao', 'Linhas de Dados'], rows, highlight=index_highlight)

    # Torna cada linha da coluna "Aba" um link clicável que leva direto para a
    # aba correspondente (referência interna do Excel: #'NomeDaAba'!A1).
    idx_ws = wb['0_Indice']
    link_font = Font(color='1D4ED8', underline='single', bold=True)
    for r in range(2, idx_ws.max_row + 1):
        cell = idx_ws.cell(row=r, column=1)
        sheet_name = cell.value
        if sheet_name and sheet_name in wb.sheetnames:
            # A forma "cell.hyperlink = '#...'" não sobrevive ao save/reload do
            # Excel (perde a location) — é preciso um Hyperlink explícito com
            # location=(sem '#') para o link interno funcionar de fato.
            cell.hyperlink = Hyperlink(ref=cell.coordinate, location=f"'{sheet_name}'!A1", display=sheet_name)
            cell.font = link_font
    idx_ws.column_dimensions['B'].width = 85

    wb.move_sheet('0_Indice', offset=-(len(wb.sheetnames) - 1))


def write_excel_workbook(summary, governance, license_rows, domain_counts, missing_email_rows, sanity_data=None, migration_data=None, identities=None, allocation_data=None, security_audit_data=None, group_baseline_data=None, role_standardization_data=None, reconciliation_data=None):
    """Creates the final consolidated governance workbook used by the pipeline."""
    wb = Workbook()
    wb.remove(wb.active)

    header_fill = PatternFill(start_color='0f172a', end_color='0f172a', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    thin_border = Border(
        left=Side(style='thin', color='e2e8f0'),
        right=Side(style='thin', color='e2e8f0'),
        top=Side(style='thin', color='e2e8f0'),
        bottom=Side(style='thin', color='e2e8f0'),
    )

    def add_sheet(title, headers, rows, highlight=None):
        """highlight(row, idx) -> hex color string (sem '#') ou None. `row` é o
        item original (dict ou lista) e `idx` é a posição (0-based) na lista de
        dados — usado para zebra-striping e para destacar linhas críticas."""
        ws = wb.create_sheet(title[:31])
        ws.append(headers)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')

        for idx, row in enumerate(rows):
            if isinstance(row, dict):
                values = []
                for h in headers:
                    v = row.get(h, '')
                    # Convert sets to strings for Excel compatibility
                    if isinstance(v, set):
                        v = '; '.join(sorted(str(x) for x in v if x)) if v else ''
                    values.append(v)
            else:
                values = row
            ws.append(values)
            if highlight:
                color = highlight(row, idx)
                if color:
                    row_fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                    for cell in ws[ws.max_row]:
                        cell.fill = row_fill

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(vertical='top', wrap_text=True)

        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = ws.dimensions
        for col in ws.columns:
            max_len = max((len(str(cell.value)) for cell in col if cell.value is not None), default=0)
            ws.column_dimensions[col[0].column_letter].width = min(max(max_len + 2, 12), 45)

    license_headers = [
        'USERID', 'DISPLAYNAME', 'EMAIL', 'DOMAIN_CATEGORY', 'MIGRATION_SCOPE',
        'ENTITLEMENT', 'LICENSE_MODEL', 'APP_POINTS', 'LOCATION_SITE',
        'USAGE_PROFILE', 'OPTIMIZATION_REC', 'OPTIMIZATION_REASON',
        'LOGIN_COUNT_90D', 'LOGIN_COUNT_60D', 'DAYS_SINCE_LAST', 'FACTOR_P50', 'FACTOR_P95',
        'FACTOR_P100', 'TITLES'
    ]

    scenario = summary.get('app_points_summary', {})
    executive_rows = [
        ['Gerado em', datetime.now().strftime('%d/%m/%Y %H:%M')],
        ['Usuarios ativos analisados', summary.get('active_profiles_count', 0)],
        ['Usuarios no plano de licenca', len(license_rows)],
        ['Usuarios sem dominio para revisao', len(missing_email_rows)],
        ['Authorized', len(scenario.get('auth_users', []))],
        ['Concurrent', len(scenario.get('conc_users', []))],
        ['Premium', len(scenario.get('premium_users', []))],
    ]
    for domain, count in sorted(domain_counts.items()):
        executive_rows.append([f'Dominio: {domain}', count])

    def executive_highlight(row, idx):
        label = str(row[0]) if row else ''
        if label.startswith('Dominio:'):
            return 'eff6ff'  # azul claro - distribuição por domínio
        if label in ('Authorized', 'Concurrent', 'Premium'):
            return 'f5f3ff'  # lilás claro - split de licença
        if label in ('Usuarios ativos analisados', 'Usuarios no plano de licenca'):
            return 'ecfdf5'  # verde claro - totais-chave
        return None

    # As abas são criadas EM ORDEM NUMÉRICA ESTRITA (1, 2, 3, 4, 5...) para que
    # a ordem visível das abas no Excel corresponda ao prefixo do nome. Antes,
    # cada bloco era adicionado fora de sequência (ex.: "7" e "8" eram criadas
    # antes de "3"), deixando as abas na ordem visual 1,2,7,8,3,4,5,6,9,9b,...
    add_sheet('1_VisaoExecutiva', ['Metrica', 'Valor'], executive_rows, highlight=executive_highlight)

    # Aba 2: MASTER — todos os dados de AD, Maximo, licença, saneamento,
    # migração e alocação numa única linha por usuário, para validação manual
    # cruzada (a antiga "aba 2 com tudo junto").
    master_rows = build_master_rows(license_rows, sanity_data, migration_data, allocation_data)

    def master_highlight(row, idx):
        if row.get('ALERTA_AD_DESATIVADO_MAXIMO_ATIVO'):
            return 'fecaca'  # vermelho - inconsistência crítica AD x Maximo
        status_ad = row.get('STATUS_AD')
        if status_ad == 'DESATIVADO':
            return 'fef3c7'  # âmbar - desativado no AD (sem alerta de Maximo ativo)
        if status_ad == 'NAO ENCONTRADO NO AD':
            return 'f1f5f9'  # cinza claro - não encontrado no AD
        if row.get('ENTITLEMENT') == 'PREMIUM':
            return 'ede9fe'  # lilás - entitlement premium
        return 'f8fafc' if idx % 2 else None  # zebra sutil no restante

    if master_rows:
        add_sheet('2_Master_Analise', list(master_rows[0].keys()), master_rows, highlight=master_highlight)

    add_sheet('3_LicenseDecisionPlan', license_headers, license_rows)

    if missing_email_rows:
        review_headers = [
            'USERID', 'DISPLAYNAME', 'STATUS', 'ENVS', 'TYPE', 'GROUPS_COUNT',
            'GROUPS', 'TITLES', 'PERSONGROUPS', 'REVIEW_REASON'
        ]
        add_sheet('4_RevisarSemDominio', review_headers, missing_email_rows)

    for sheet_name, key in [
        ('5_ReusoUSERID_CrossEnv', 'cross_env'),
        ('6_ConflitosLoginID', 'login_conflicts'),
        ('7_FilaSaneamento', 'worklist'),
    ]:
        rows = governance.get(key, [])
        if rows:
            add_sheet(sheet_name, list(rows[0].keys()), rows)

    # Adicionar Abas 10-17: Saneamento de Identidades (AD vs Maximo)
    if sanity_data:
        add_sanity_sheets(wb, sanity_data, add_sheet)

    # Adicionar Abas 18-19: Recomendações de Migração
    if migration_data:
        add_migration_sheets(wb, migration_data, add_sheet)

    # Adicionar Abas 20-21: Saneamento de Alocação (Maximo 9)
    if allocation_data:
        add_allocation_sheets(wb, allocation_data, add_sheet)

    # Adicionar Aba 22: Usuários Ativos Únicos do Maximo
    if identities:
        add_maximo_active_users_sheet(wb, identities, add_sheet)

    # Adicionar Aba 23: Acessos por Tipo de Perfil
    if identities:
        add_profile_access_sheet(wb, identities, add_sheet)

    # Adicionar Aba 16: Auditoria de Acesso
    if identities:
        add_audit_sheet(wb, identities, governance.get('groupuser', []), governance.get('logintrack', []),
                         governance.get('maxuserstatus', []), add_sheet)

    # Aba 25: resumo executivo consolidado (migração + SoD + cenário conciliado)
    if migration_data or security_audit_data or reconciliation_data:
        add_consolidated_summary(wb, migration_data, security_audit_data, reconciliation_data, add_sheet)

    # Abas 26-28: Auditoria de Segregação de Funções (grupos, pessoas, evidências)
    if security_audit_data:
        add_security_audit_sheets(wb, security_audit_data, add_sheet)

    # Abas 29-31: Modelo de acesso por cargo (baseline+padrão, desvios, grupos duplicados)
    if group_baseline_data or role_standardization_data:
        add_access_model_sheets(wb, group_baseline_data, role_standardization_data, add_sheet)

    # Aba 32: Cenário Conciliado — detalhe por usuário
    if reconciliation_data:
        add_reconciliation_sheets(wb, reconciliation_data, add_sheet)

    # Aba 0 (Índice): navegação com descrição e contagem de linhas de cada
    # aba — necessária num workbook com mais de 20 abas numeradas.
    add_index_sheet(wb, add_sheet)

    out_path = OUT_DIR / 'maximo_risk_and_optimization_workbook.xlsx'
    try:
        # Tentar remover arquivo existente se houver
        if out_path.exists():
            out_path.unlink()
        wb.save(out_path)
        print(f'✓ WROTE {out_path.name}')
    except PermissionError:
        # Se não conseguir sobrescrever, usar nome com timestamp
        alt_path = OUT_DIR / f'maximo_risk_and_optimization_workbook_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        wb.save(alt_path)
        print(f'✓ WROTE {alt_path.name} (arquivo original estava aberto)')

# --- Sanity Excel Sheets ---
def add_sanity_sheets(wb, sanity_data, add_sheet):
    """Adiciona abas do Excel com dados de saneamento de identidades."""

    # Aba 10: Resumo de Saneamento
    stats = sanity_data['stats']
    SECTION_LABELS = {
        'MATCH POR EMAIL', 'DIVERGÊNCIAS', 'MATCH POR PREFIXO (USERID)',
        'MAXIMO SEM EMAIL', 'AUDITORIA CRÍTICA',
    }
    summary_rows = [
        ['Total AD', stats['total_ad']],
        ['Total Maximo (identities)', stats['total_maximo_identities']],
        ['Total Maximo (USERIDs únicos)', stats['total_maximo_userids']],
        ['', ''],
        ['MATCH POR EMAIL', ''],
        ['Match perfeito', stats['match_email']],
        ['Apenas no AD', stats['only_ad']],
        ['Apenas no Maximo', stats['only_maximo']],
        ['', ''],
        ['DIVERGÊNCIAS', ''],
        ['Nomes diferentes (mesmo email)', stats['name_divergences']],
        ['Múltiplos USERIDs (mesmo email)', stats['multi_userid']],
        ['Domínios divergentes', stats['domain_divergences']],
        ['', ''],
        ['MATCH POR PREFIXO (USERID)', ''],
        ['Match encontrado', stats['prefix_match']],
        ['Sem match no Maximo', stats['no_match']],
        ['', ''],
        ['MAXIMO SEM EMAIL', ''],
        ['Com match no AD', stats['maximo_sem_email_match']],
        ['Sem match no AD', stats['maximo_sem_email_nomatch']],
        ['', ''],
        ['AUDITORIA CRÍTICA', ''],
        ['Desativados no AD mas ativos no Maximo (aba 10, ACHADO=AD_DESATIVADO_MAS_ATIVO)', stats['ad_disabled_ativos_maximo']],
    ]

    def resumo_highlight(row, idx):
        label = str(row[0]) if row else ''
        if label in SECTION_LABELS:
            return 'dbeafe'  # azul claro - separador de seção
        if 'AD mas ativos no Maximo' in label:
            return 'fecaca'  # vermelho claro - achado crítico
        return None

    add_sheet('9_Saneamento_Resumo', ['Métrica', 'Valor'], summary_rows, highlight=resumo_highlight)

    # Aba 10: TODOS os achados do saneamento AD × Maximo numa só aba
    # (consolidação a pedido do usuário 2026-07-11 — eram 7 abas, 11-17).
    # Filtre a coluna ACHADO para ver cada categoria; o achado crítico
    # (AD_DESATIVADO_MAS_ATIVO) fica destacado em vermelho.
    achados = []

    def _add(achado, email, nome_ad, userids, nomes_mx, envs, status_mx, ad_on, grupos_ad, detalhe):
        achados.append({
            'ACHADO': achado, 'EMAIL_AD': email, 'NOME_AD': nome_ad,
            'USERIDS_MAXIMO': userids, 'NOMES_MAXIMO': nomes_mx, 'AMBIENTES': envs,
            'STATUS_MAXIMO': status_mx, 'AD_HABILITADO': ad_on,
            'QTD_GRUPOS_AD': grupos_ad, 'DETALHE': detalhe,
        })

    for d in sanity_data['analises'].get('ad_disabled_ativos_maximo', []):
        _add('AD_DESATIVADO_MAS_ATIVO', d['email'], d['ad_displayname'], d['maximo_userids'],
             '', d['maximo_envs'], d['maximo_statuses'], 'Não', d['ad_groups_count'],
             f"match {d.get('match_type', 'EMAIL')}; ambientes ativos/total: {d.get('qtd_envs_ativos_de_total', '')}")
    for d in sanity_data['analises'].get('name_divergences', []):
        _add('DIVERGENCIA_NOME', d['email'], d['ad_displayname'], d['maximo_userids'],
             d['maximo_names'], d['maximo_envs'], d['maximo_statuses'],
             'Sim' if d['ad_enabled'] else 'Não', d['ad_groups_count'], '')
    for d in sanity_data['analises'].get('multi_userid', []):
        _add('MULTIPLOS_USERIDS', d['email'], d['ad_displayname'], d['userids'],
             '', d['envs'], d['statuses'], '', '', f"{d['qtd_userids']} USERIDs para o mesmo e-mail")
    for d in sanity_data['analises'].get('prefix_match', []):
        _add('MATCH_POR_PREFIXO', d['email'], d['ad_displayname'], d['maximo_userid'],
             d['maximo_displaynames'], d['maximo_envs'], d['maximo_statuses'],
             'Sim' if d['ad_enabled'] else 'Não', d['ad_groups_count'],
             f"emails Maximo: {d['maximo_emails']}")
    for d in sanity_data['analises'].get('no_match', []):
        _add('AD_SEM_MATCH_MAXIMO', d['email'], d['ad_displayname'], '', '', '', '',
             'Sim' if d['ad_enabled'] else 'Não', d['ad_groups_count'], f"prefixo: {d['prefix']}")
    for d in sanity_data['analises'].get('maximo_sem_email_match', []):
        _add('MAXIMO_SEM_EMAIL_COM_AD', d['ad_email'], d['ad_displayname'], d['userid'],
             d['maximo_displaynames'], d['maximo_envs'], d['maximo_statuses'],
             'Sim' if d['ad_enabled'] else 'Não', d['ad_groups_count'],
             f"títulos: {d['maximo_titles']}")
    for d in sanity_data['analises'].get('domain_divergences', []):
        _add('DIVERGENCIA_DOMINIO', d['email'], d['ad_displayname'], d['maximo_userid'],
             '', d['maximo_env'], d['maximo_status'], '', '',
             f"AD {d['ad_domain']} vs Maximo {d['maximo_domain']}")

    if achados:
        headers = ['ACHADO', 'EMAIL_AD', 'NOME_AD', 'USERIDS_MAXIMO', 'NOMES_MAXIMO',
                   'AMBIENTES', 'STATUS_MAXIMO', 'AD_HABILITADO', 'QTD_GRUPOS_AD', 'DETALHE']
        add_sheet('10_Saneamento_AD_Achados', headers, achados,
                  highlight=lambda r, i: 'fee2e2' if r.get('ACHADO') == 'AD_DESATIVADO_MAS_ATIVO' else None)


# --- Maximo Active Users Excel Sheet ---
def add_maximo_active_users_sheet(wb, identities, add_sheet):
    """Adiciona aba com todos os usuários únicos ativos no Maximo."""

    def clean_value(v):
        """Remove caracteres ilegais do Excel (caracteres de controle)."""
        if v is None:
            return ''
        s = str(v)
        return ''.join(c if ord(c) >= 32 or c in '\n\r\t' else '' for c in s)

    # Filtrar apenas usuários ativos
    active_users = [r for r in identities if r.get('STATUS', '').strip().upper() == 'ACTIVE']

    # Deduplicar por (ENV_DB, USERID) — não só USERID. A mesma conta pode
    # estar ativa em múltiplos ambientes (ex.: HELPDESK, ITEAM em 6 bases) e
    # a aba exibe ENV_DB por linha; dedup só por USERID descartava essas
    # ocorrências (auditoria 2026-07-13 confirmou 210/1840 linhas ativas
    # perdidas, 73 USERIDs afetados).
    seen_keys = set()
    unique_active = []
    for user in active_users:
        userid = user.get('USERID', '').strip().upper()
        env = user.get('ENV_DB', '').strip().upper()
        key = (env, userid)
        if userid and key not in seen_keys:
            seen_keys.add(key)
            unique_active.append(user)

    # Preparar headers e rows
    headers = ['USERID', 'DISPLAYNAME', 'PRIMARYEMAIL', 'ENV_DB', 'STATUS',
               'TYPE', 'DEFSITE', 'FIRSTNAME', 'LASTNAME', 'TITLE', 'PERSONGROUP']

    rows = []
    for user in unique_active:
        rows.append({
            'USERID': clean_value(user.get('USERID', '')),
            'DISPLAYNAME': clean_value(user.get('DISPLAYNAME', '')),
            'PRIMARYEMAIL': clean_value(user.get('PRIMARYEMAIL', '')),
            'ENV_DB': clean_value(user.get('ENV_DB', '')),
            'STATUS': clean_value(user.get('STATUS', '')),
            'TYPE': clean_value(user.get('TYPE', '')),
            'DEFSITE': clean_value(user.get('DEFSITE', '')),
            'FIRSTNAME': clean_value(user.get('FIRSTNAME', '')),
            'LASTNAME': clean_value(user.get('LASTNAME', '')),
            'TITLE': clean_value(user.get('TITLE', '')),
            'PERSONGROUP': clean_value(user.get('PERSONGROUP', '')),
        })

    add_sheet('14_Maximo_Usuarios_Ativos', headers, rows)
    print(f'✓ Aba 14 adicionada: {len(rows)} usuários ativos únicos do Maximo')


# --- Profile Access Excel Sheet ---
def add_profile_access_sheet(wb, identities, add_sheet):
    """Adiciona aba com detalhamento de acessos por tipo de perfil."""

    def clean_value(v):
        """Remove caracteres ilegais do Excel (caracteres de controle)."""
        if v is None:
            return ''
        s = str(v)
        return ''.join(c if ord(c) >= 32 or c in '\n\r\t' else '' for c in s)

    # Filtrar apenas usuários ativos
    active_users = [r for r in identities if r.get('STATUS', '').strip().upper() == 'ACTIVE']

    # Agrupar por tipo
    by_type = {}
    for user in active_users:
        t = user.get('TYPE', '').strip() or 'N/A'
        if t not in by_type:
            by_type[t] = []
        by_type[t].append(user)

    # Preparar headers
    headers = ['Tipo', 'Qtd Usuários', 'USERIDs Únicos', 'Ambientes', 'Títulos', 'PersonGroups']

    rows = []
    for tipo, users in sorted(by_type.items()):
        userids = set()
        envs = set()
        titulos = set()
        persongroups = set()

        for user in users:
            userids.add(clean_value(user.get('USERID', '')))
            envs.add(clean_value(user.get('ENV_DB', '')))
            titulos.add(clean_value(user.get('TITLE', '')))
            persongroups.add(clean_value(user.get('PERSONGROUP', '')))

        rows.append({
            'Tipo': clean_value(tipo),
            'Qtd Usuários': len(users),
            'USERIDs Únicos': len(userids),
            'Ambientes': '; '.join(sorted(e for e in envs if e)),
            'Títulos': '; '.join(sorted(t for t in titulos if t))[:100],
            'PersonGroups': '; '.join(sorted(pg for pg in persongroups if pg))[:100],
        })

    add_sheet('15_Acessos_por_Perfil', headers, rows)
    print(f'✓ Aba 15 adicionada: {len(rows)} tipos de perfil analisados')


# --- Audit Excel Sheet ---
# --- Security Audit (Emissor x Aprovador) Excel Sheets ---
def add_consolidated_summary(wb, migration_data, security_audit_data, reconciliation_data, add_sheet):
    """Aba única de resumo executivo — funde os antigos resumos de Migração
    (18), SoD (25) e Cenário Conciliado (38) em seções de uma só aba
    (pedido do usuário 2026-07-11: menos abas para facilitar a análise)."""
    rows = []

    if migration_data:
        rows.append(['MIGRAÇÃO MAS 9 — recomendações por tipo (detalhe: aba 11)', ''])
        tipo_counts = {}
        for r in migration_data:
            tipo_counts[r['tipo']] = tipo_counts.get(r['tipo'], 0) + 1
        for tipo, count in sorted(tipo_counts.items(), key=lambda x: -x[1]):
            rows.append([f'  {tipo}', count])
        rows.append(['', ''])

    if reconciliation_data:
        s = reconciliation_data['stats']
        nc, nr = s['nem_conciliado'], s['nem_realista']
        rows += [
            ['CENÁRIO CONCILIADO — dimensionamento oficial MAS 9 (detalhe: aba 24)', ''],
            ['Maximo ativos / conciliados com AD ativo', f"{s['maximo_ativos_total']} / {s['conciliados']}"],
            ['  ...match por e-mail / prefixo / nome', f"{s['conciliados_por_email']} / {s['conciliados_por_prefixo']} / {s['conciliados_por_nome']}"],
            ['Licença estatística: AUTHORIZED / CONCURRENT', f"{s['conciliados_authorized']} / {s['conciliados_concurrent']}"],
            ['Reserva fixa Authorized (AppPoints)', s['reserva_authorized']],
            ['Terceiros ativos mantidos como Concurrent', s['terceiros_ativos']],
            ['Não conciliados sem uso em 90d (limpeza, não migram)', s['nao_conciliados_sem_uso']],
            ['NEM só-conciliados — P50 / P95 / P100', f"{nc['p50']} / {nc['p95']} / {nc['p100']}"],
            ['NEM realista (+terceiros) — P50 / P95 / P100 (teto 1.200)', f"{nr['p50']} / {nr['p95']} / {nr['p100']}"],
            ['', ''],
        ]

    if security_audit_data:
        stats = security_audit_data['stats']
        rows += [
            ['SEGREGAÇÃO DE FUNÇÕES (SoD) EM COMPRAS — detalhe: abas 18-20', ''],
            ['Ambientes cobertos', ', '.join(stats['envs_covered'])],
            ['Grupos estruturalmente conflitantes (Nível 1)', stats['total_group_conflicts']],
            ['Pessoas únicas ATIVAS com conflito (Nível 2)', stats['distinct_users_active']],
        ]
        for app, qtd in stats['distinct_users_active_by_app'].items():
            rows.append([f'  ...ativas em {app}', qtd])
        rows += [
            ['Usuários com acesso MAXADMIN', stats['total_maxadmin_users']],
            ['EVIDÊNCIA: mesma pessoa submeteu E aprovou (365d)', stats['total_real_evidence_cases']],
            ['  ...valor envolvido (USD, por PR única)', f"{stats['total_real_evidence_value']:,.2f}"],
            ['  ...CRÍTICO (2ª instância exigida e não houve)', stats['total_critical_evidence_cases']],
            ['  ...valor dos CRÍTICOS (USD)', f"{stats['total_critical_evidence_value']:,.2f}"],
            ['AUTOAPROVAÇÃO DIRETA (solicitante real = aprovador)', stats['total_self_approval_cases']],
            ['  ...valor envolvido (USD)', f"{stats['total_self_approval_value']:,.2f}"],
            ['CADEIA PR→PO (aprovou a PR e gerou a PO)', stats['total_pr_po_chain_cases']],
        ]

    def resumo_highlight(row, idx):
        label = str(row[0])
        if label and not label.startswith('  ') and label.rstrip('') == label and ('—' in label or label.isupper()):
            return 'dbeafe'
        if 'CRÍTICO' in label or 'AUTOAPROVAÇÃO' in label or 'ATIVAS com conflito' in label:
            return 'fecaca'
        return None

    add_sheet('17_Resumo_Consolidado', ['Métrica', 'Valor'], rows, highlight=resumo_highlight)


def add_security_audit_sheets(wb, security_audit_data, add_sheet):
    """Abas da auditoria de segregação de funções (SoD) em Compras,
    consolidadas em 3 abas (era 9): grupos (26), pessoas incl. MAXADMIN (27)
    e evidências das 3 camadas numa só aba com coluna CAMADA (28). Os
    antigos recortes de PO (31/32) eram subconjuntos filtráveis pela coluna
    APP e foram removidos."""

    # Aba 26: Nível 1 — grupos estruturalmente conflitantes (todas as apps;
    # filtre a coluna APP para ver só PO/PR/Requisição Simplificada)
    if security_audit_data['group_conflicts']:
        headers = ['ENVIRONMENT', 'GROUPNAME', 'APP', 'APP_LABEL', 'DESCRIPTION',
                   'OPCOES_EMISSOR', 'OPCOES_APROVADOR', 'RECOMENDACAO']
        add_sheet('18_SoD_Grupos', headers, security_audit_data['group_conflicts'],
                  highlight=lambda r, i: 'fee2e2')

    # Aba 27: pessoas — conflitos de Nível 2 + membros de MAXADMIN, com
    # coluna TIPO para filtrar
    pessoas = []
    for c in security_audit_data.get('user_conflicts', []):
        pessoas.append({**c, 'TIPO': 'CONFLITO_SOD'})
    for m in security_audit_data.get('maxadmin_users', []):
        pessoas.append({'TIPO': 'MAXADMIN', 'ENVIRONMENT': m['ENVIRONMENT'], 'USERID': m['USERID'],
                        'DISPLAYNAME': m['DISPLAYNAME'], 'TITLE': m['TITLE'], 'STATUS': m['STATUS']})
    if pessoas:
        headers = ['TIPO', 'ENVIRONMENT', 'USERID', 'DISPLAYNAME', 'TITLE', 'STATUS', 'APP', 'APP_LABEL',
                   'GRUPOS_EMISSOR', 'GRUPOS_APROVADOR', 'ORIGEM_CONFLITO', 'RECOMENDACAO',
                   'PAPEL_RECOMENDADO', 'JUSTIFICATIVA_PAPEL']

        def user_highlight(row, idx):
            if row.get('TIPO') == 'MAXADMIN':
                return 'e2e8f0'
            if row.get('STATUS', '').strip().upper() not in ('ACTIVE', 'ATIVO', 'ENABLED'):
                return 'f1f5f9'
            return 'fee2e2' if row.get('ORIGEM_CONFLITO') == 'MESMO_GRUPO' else 'fef3c7'

        add_sheet('19_SoD_Pessoas', headers, pessoas, highlight=user_highlight)

    # Aba 28: evidências das 3 camadas numa só aba (coluna CAMADA):
    #   SUBMETEU_E_APROVOU  — WFTRANSACTION WAPPR+APPR pela mesma pessoa
    #   AUTOAPROVACAO       — solicitante real (OOG_REQUESTEDBY) = aprovador
    #   APROVOU_PR_GEROU_PO — cadeia PR→PO (OOG_CREAPOGRP)
    evid = []
    for e in security_audit_data.get('real_evidence', []):
        evid.append({
            'CAMADA': 'SUBMETEU_E_APROVOU', 'SEVERIDADE': e['SEVERIDADE'],
            'SITEID': e['SITEID'], 'PRNUM': e['PRNUM'], 'DESCRIPTION': e['DESCRIPTION'],
            'TOTALCOST_USD': e['TOTALCOST'], 'STATUS_DOC': e['STATUS'],
            'PESSOA': e['PERSONID'], 'NOME_PESSOA': e.get('NOME_PESSOA', ''),
            'TITULO_PESSOA': e.get('TITULO_PESSOA', ''), 'STATUS_PESSOA': e.get('STATUS_PESSOA', ''),
            'DATA_SUBMISSAO': e.get('DATA_SUBMISSAO', ''), 'DATA_APROVACAO': e.get('DATA_APROVACAO', ''),
            'ROTEADO_2A_INSTANCIA': e.get('ROTEADO_2A_INSTANCIA', ''), 'PONUM_GERADA': '',
        })
    for e in security_audit_data.get('self_approval_evidence', []):
        evid.append({
            'CAMADA': 'AUTOAPROVACAO', 'SEVERIDADE': e['SEVERIDADE'],
            'SITEID': e['SITEID'], 'PRNUM': e['PRNUM'], 'DESCRIPTION': e['DESCRIPTION'],
            'TOTALCOST_USD': e['TOTALCOST'], 'STATUS_DOC': e['STATUS'],
            'PESSOA': e['SOLICITANTE_REAL'], 'NOME_PESSOA': e.get('NOME_PESSOA', ''),
            'TITULO_PESSOA': e.get('TITULO_PESSOA', ''), 'STATUS_PESSOA': e.get('STATUS_PESSOA', ''),
            'DATA_SUBMISSAO': '', 'DATA_APROVACAO': e.get('DATA_APROVACAO', ''),
            'ROTEADO_2A_INSTANCIA': e.get('ROTEADO_2A_INSTANCIA', ''), 'PONUM_GERADA': '',
        })
    for e in security_audit_data.get('pr_po_chain_evidence', []):
        evid.append({
            'CAMADA': 'APROVOU_PR_GEROU_PO', 'SEVERIDADE': 'CRITICO',
            'SITEID': e['SITEID'], 'PRNUM': e['PRNUM'], 'DESCRIPTION': e['DESCRIPTION'],
            'TOTALCOST_USD': e['TOTALCOST'], 'STATUS_DOC': e['STATUS'],
            'PESSOA': e['PERSONID'], 'NOME_PESSOA': e.get('NOME_PESSOA', ''),
            'TITULO_PESSOA': e.get('TITULO_PESSOA', ''), 'STATUS_PESSOA': e.get('STATUS_PESSOA', ''),
            'DATA_SUBMISSAO': e.get('DATA_APROVACAO_PR', ''), 'DATA_APROVACAO': e.get('DATA_CRIACAO_PO', ''),
            'ROTEADO_2A_INSTANCIA': '', 'PONUM_GERADA': e.get('PONUM_GERADA', ''),
        })
    if evid:
        headers = ['CAMADA', 'SEVERIDADE', 'SITEID', 'PRNUM', 'DESCRIPTION', 'TOTALCOST_USD',
                   'STATUS_DOC', 'PESSOA', 'NOME_PESSOA', 'TITULO_PESSOA', 'STATUS_PESSOA',
                   'DATA_SUBMISSAO', 'DATA_APROVACAO', 'ROTEADO_2A_INSTANCIA', 'PONUM_GERADA']

        def evid_highlight(row, idx):
            if row.get('CAMADA') == 'AUTOAPROVACAO' or row.get('SEVERIDADE') == 'CRITICO':
                return 'fecaca'
            return None

        add_sheet('20_SoD_Evidencias', headers, evid, highlight=evid_highlight)


def add_access_model_sheets(wb, group_baseline_data, role_standardization_data, add_sheet):
    """Abas do modelo de acesso por cargo, consolidadas em 3 (era 4):
    - 29_Cargo_x_Grupos: funde o baseline por cargo/unidade (ex-35) e o
      padrão global recomendado para o MAS 9 (ex-37), com coluna UNIDADE
      ('TODAS (padrão MAS 9)' para o alvo global).
    - 30_Perfil_Cargo_Desvios: pessoas fora do padrão do próprio cargo.
    - 31_Grupos_Duplicados: clusters de grupos com permissão quase idêntica
      sob nomes diferentes (material para a terceirizada fundir)."""
    merged = []
    for t in (role_standardization_data or {}).get('role_targets', []):
        merged.append({
            'CARGO': t['CARGO_NORMALIZADO'],
            'UNIDADE': 'TODAS (padrão MAS 9)',
            'QTD_PESSOAS': t['QTD_PESSOAS'],
            'GRUPOS_PADRAO': t['GRUPO_PADRAO_RECOMENDADO'],
            'CONSISTENTE_ENTRE_UNIDADES': t['CONSISTENTE_ENTRE_UNIDADES'],
            'ACAO': t['ACAO'],
        })
    for p in (group_baseline_data or {}).get('profile_rows', []):
        merged.append({
            'CARGO': p['TITLE'],
            'UNIDADE': p['ENVIRONMENT'],
            'QTD_PESSOAS': p['QTD_PESSOAS'],
            'GRUPOS_PADRAO': p['GRUPOS_PADRAO'],
            'CONSISTENTE_ENTRE_UNIDADES': '',
            'ACAO': '',
        })
    if merged:
        merged.sort(key=lambda x: (x['CARGO'], x['UNIDADE'] != 'TODAS (padrão MAS 9)', x['UNIDADE']))
        headers = ['CARGO', 'UNIDADE', 'QTD_PESSOAS', 'GRUPOS_PADRAO',
                   'CONSISTENTE_ENTRE_UNIDADES', 'ACAO']
        add_sheet('21_Cargo_x_Grupos', headers, merged,
                  highlight=lambda r, i: 'fecaca' if r.get('CONSISTENTE_ENTRE_UNIDADES') is False
                  else ('dbeafe' if r.get('UNIDADE') == 'TODAS (padrão MAS 9)' else None))

    deviation_rows = (group_baseline_data or {}).get('deviation_rows', [])
    if deviation_rows:
        headers = ['ENVIRONMENT', 'USERID', 'DISPLAYNAME', 'TITLE', 'COHORT_SIZE',
                   'GRUPOS_EXCESSO', 'GRUPOS_FALTANTES', 'QTD_EXCESSO', 'QTD_FALTANTES']
        add_sheet('22_Perfil_Cargo_Desvios', headers, deviation_rows,
                  highlight=lambda r, i: 'fecaca' if r.get('QTD_EXCESSO', 0) > 0 else 'fef3c7')

    duplicate_clusters = (role_standardization_data or {}).get('duplicate_group_clusters', [])
    if duplicate_clusters:
        headers = ['CANONICO', 'MEMBROS', 'QTD_MEMBROS', 'DESCRICOES',
                   'ALERTA_PRIVILEGIO_DIFERENTE', 'ALERTA_NOMES_DIVERGENTES']
        rows = [{**c, 'MEMBROS': '; '.join(c['MEMBROS'])} for c in duplicate_clusters]
        add_sheet('23_Grupos_Duplicados', headers, rows,
                  highlight=lambda r, i: 'fecaca' if r.get('ALERTA_PRIVILEGIO_DIFERENTE') else 'fef3c7')


def add_reconciliation_sheets(wb, reconciliation_data, add_sheet):
    """Aba de detalhe por usuário do Cenário Conciliado (o resumo foi
    fundido na 25_Resumo_Consolidado)."""
    rows = reconciliation_data.get('rows', [])
    if rows:
        headers = list(rows[0].keys())
        add_sheet('24_Cenario_Conciliado_Usuarios', headers, rows,
                  highlight=lambda r, i: 'fef3c7' if r.get('POPULACAO') == 'TERCEIRO_ATIVO' else None)


def add_audit_sheet(wb, identities, groupuser, logintrack, maxuserstatus, add_sheet):
    """Adiciona aba única de auditoria (ativos + inativados) — pedido de
    auditoria 2026-07-13.

    TYPE vem de MAXUSER (via consolidated_user_identity.csv) — valores reais
    'TYPE 1'..'TYPE 10' — não de PERSONGROUPVIEW.employeetype (campo
    diferente, que ficava vazio/errado aqui). DATA_CONCESSAO/DATA_INATIVACAO
    vêm de MAXUSERSTATUS (histórico real de mudança de status da CONTA —
    mesma fonte da tela "View History" do Maximo), NÃO de
    PERSONGROUPVIEW.statusdate: esse campo rastreia o status da PESSOA, não
    da conta de login, e usá-lo como "data de inativação" produzia datas
    erradas (ex.: USERID AAJUNIOR mostrava 21/04 — na verdade a data de
    CRIAÇÃO da pessoa — quando a conta só foi inativada de fato em 24/06,
    confirmado via MAXUSERSTATUS). MAXUSERSTATUS não é replicado entre
    ambientes (cada ambiente tem seu próprio histórico), por isso o lookup é
    por (ambiente, USERID). QTD_MUDANCAS_STATUS mostra quantas transições
    ACTIVE/INACTIVE aquela conta já teve — uma conta com várias reativações
    é um sinal de que o "tempo sem acesso" não é linear (foi zerado a cada
    reativação manual), então DIAS_SEM_ACESSO por si só pode subestimar o
    problema ou, ao contrário, ignorar uma reativação recente legítima.
    DEFSITE,
    BASE_DO_PERFIL e BASE_LOGADA passam por norm_env() para exibição (ex.:
    'OP-BASE'/'BASE-UNP' -> 'BASE', 'NORBE09' -> 'N09') — mesmos ambientes,
    nomes diferentes por fonte. GRUPOS_ACESSO é a lista de GROUPNAME
    (GROUPUSER) do USERID no mesmo ambiente. DATA_ULTIMO_ACESSO e
    BASE_LOGADA vêm do histórico de LOGINTRACKING (evento real de acesso,
    não a base de cadastro). ALERTA_INATIVIDADE sinaliza conta ATIVA sem
    login há mais de 50 dias (ou nunca logada) — pedido de auditoria
    2026-07-13: "conta ativa sem acesso há mais de 50 dias deve estar
    inativa". Referência de "hoje" é a data mais recente encontrada no
    próprio LOGINTRACKING (não a data corrida do sistema).

    A aba inclui TODA identidade (ativa ou inativa), mesmo sem registro em
    PERSONGROUPVIEW — quando não há data conhecida, DATA_CONCESSAO e
    DATA_INATIVACAO ficam em branco em vez de a pessoa desaparecer da lista
    sem aviso (686/1900 ativos, 36%, incluindo contas MAXADMIN, ficavam de
    fora antes desta correção — auditoria 2026-07-13)."""

    def clean_value(v):
        """Remove caracteres ilegais do Excel (caracteres de controle)."""
        if v is None:
            return ''
        s = str(v)
        return ''.join(c if ord(c) >= 32 or c in '\n\r\t' else '' for c in s)

    def parse_dt(s):
        for fmt in ("%Y-%m-%d-%H.%M.%S.%f", "%Y-%m-%d %H:%M:%S.%f", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d-%H.%M.%S", "%Y-%m-%d"):
            try:
                return datetime.strptime(s.strip(), fmt)
            except (ValueError, AttributeError):
                continue
        return None

    # ENV_DB (identities) usa nomes como 'NORBE09'; GROUPUSER/LOGINTRACKING
    # usam o alias curto 'N09' — sem normalizar, o cruzamento por ambiente
    # falha silenciosamente e GRUPOS_ACESSO fica vazio para essas contas.
    ENV_ALIAS = {'NORBE06': 'N06', 'NORBE08': 'N08', 'NORBE09': 'N09',
                 'BASE-UNP': 'BASE', 'OP-BASE': 'BASE', 'ODRL-SP': 'BASE'}

    def norm_env(env):
        e = (env or '').strip().upper()
        return ENV_ALIAS.get(e, e)

    # Histórico real de mudança de STATUS da CONTA (não da pessoa), por
    # (ambiente normalizado, USERID) — ver docstring sobre por que isto
    # substitui PERSONGROUPVIEW.statusdate. MAXUSERSTATUS não é replicado
    # entre ambientes, então cada (env, uid) só existe no ambiente real da
    # conta.
    status_changes_by_env_userid = defaultdict(list)
    for r in maxuserstatus:
        env = norm_env(r.get('ENVIRONMENT', ''))
        uid = r.get('USERID', '').strip().upper()
        changedate = r.get('CHANGEDATE', '').strip()
        if not uid or not changedate:
            continue
        dt = parse_dt(changedate)
        if not dt:
            continue
        status_changes_by_env_userid[(env, uid)].append((dt, changedate))

    latest_status_change_by_env_userid = {}
    qtd_mudancas_by_env_userid = {}
    for key, changes in status_changes_by_env_userid.items():
        changes.sort(key=lambda c: c[0])
        latest_status_change_by_env_userid[key] = changes[-1][1]  # CHANGEDATE mais recente (texto original)
        qtd_mudancas_by_env_userid[key] = len(changes)

    # Grupos de acesso (GROUPNAME) por (ENVIRONMENT normalizado, USERID).
    groups_by_env_userid = defaultdict(set)
    for g in groupuser:
        env = norm_env(g.get('ENVIRONMENT', ''))
        uid = g.get('USERID', '').strip().upper()
        grp = g.get('GROUPNAME', '').strip()
        if uid and grp:
            groups_by_env_userid[(env, uid)].add(grp)

    # Último login real (data + base logada) por (ambiente normalizado,
    # USERID), via LOGINTRACKING — não só por USERID: a mesma pessoa pode
    # ter conta ativa num ambiente e inativa em outro (MAXUSERSTATUS não é
    # replicado — ver docstring), então comparar "dias sem acesso" ou
    # "login depois da inativação" precisa olhar o login DO MESMO ambiente
    # da linha, não o login mais recente da pessoa em QUALQUER ambiente
    # (auditoria 2026-07-13: sem isso, ~1100 contas legitimamente inativas
    # num ambiente pareciam "logaram depois de inativadas" só porque a
    # pessoa usou uma conta ATIVA em outro ambiente).
    # max_login_dt é usado como referência de "hoje" para calcular dias sem
    # acesso — mesmo padrão de allocation_analyzer.py (data corrida do
    # servidor não é confiável neste ambiente de execução).
    last_login_by_env_userid = {}
    max_login_dt = None
    for rec in logintrack:
        if (rec.get('ATTEMPTRESULT') or '').strip().upper() != 'LOGIN':
            continue
        uid = (rec.get('USERID') or '').strip().upper()
        if not uid:
            continue
        rec_env = norm_env(rec.get('ENVIRONMENT', ''))
        dt = parse_dt(rec.get('ATTEMPTDATE', ''))
        if not dt:
            continue
        if not max_login_dt or dt > max_login_dt:
            max_login_dt = dt
        key = (rec_env, uid)
        prev = last_login_by_env_userid.get(key)
        if not prev or dt > prev[0]:
            last_login_by_env_userid[key] = (dt, rec_env)
    INATIVIDADE_LIMITE_DIAS = 50

    # Base: TODAS as identidades (ativos e inativos). Antes, quem não tinha
    # registro em PERSONGROUPVIEW (686/1900 ativos, 36% — auditoria
    # 2026-07-13, incluindo contas MAXADMIN) era excluído da aba sem nenhum
    # aviso. Agora aparece igual, só com DATA_CONCESSAO/DATA_INATIVACAO em
    # branco quando não há histórico em MAXUSERSTATUS para aquele (ambiente,
    # USERID).
    headers = ['USERID', 'DISPLAYNAME', 'EMAIL', 'STATUS', 'DATA_CONCESSAO', 'DATA_INATIVACAO', 'QTD_MUDANCAS_STATUS',
               'TYPE', 'DEFSITE', 'TITLE', 'PERSONGROUP', 'BASE_DO_PERFIL', 'GRUPOS_ACESSO', 'DATA_ULTIMO_ACESSO',
               'BASE_LOGADA', 'DIAS_SEM_ACESSO', 'ALERTA_INATIVIDADE']

    rows = []
    for r in identities:  # Sem limite de linhas — ver nota acima sobre não esconder ninguém
        env = r.get('ENV_DB', '').strip()
        uid = r.get('USERID', '').strip().upper()
        is_active = r.get('STATUS', '').strip().upper() == 'ACTIVE'
        env_uid_key = (norm_env(env), uid)
        grupos = sorted(groups_by_env_userid.get(env_uid_key, []))
        last_login = last_login_by_env_userid.get(env_uid_key)
        statusdate = latest_status_change_by_env_userid.get(env_uid_key, '')
        qtd_mudancas = qtd_mudancas_by_env_userid.get(env_uid_key, 0)

        # Conta ATIVA sem login há mais de 50 dias (ou nunca logou) é
        # sinalizada explicitamente — pedido de auditoria 2026-07-13:
        # "se tiver alguma conta ativa sem acesso há mais de 50 dias,
        # apontar na aba de auditoria".
        dias_sem_acesso = ''
        alerta = ''
        if is_active and max_login_dt:
            if last_login:
                dias_sem_acesso = (max_login_dt - last_login[0]).days
                if dias_sem_acesso > INATIVIDADE_LIMITE_DIAS:
                    alerta = f'REVISAR - {dias_sem_acesso}d sem acesso'
            else:
                alerta = 'REVISAR - nenhum login registrado'

        rows.append({
            'USERID': clean_value(r.get('USERID', '')),
            'DISPLAYNAME': clean_value(r.get('DISPLAYNAME', '')),
            'EMAIL': clean_value(r.get('PRIMARYEMAIL', '')),
            'STATUS': clean_value(r.get('STATUS', '')),
            'DATA_CONCESSAO': clean_value(statusdate) if is_active else '',
            'DATA_INATIVACAO': clean_value(statusdate) if not is_active else '',
            'QTD_MUDANCAS_STATUS': qtd_mudancas,
            'TYPE': clean_value(r.get('TYPE', '')),
            'DEFSITE': clean_value(norm_env(r.get('DEFSITE', ''))),
            'TITLE': clean_value(r.get('TITLE', '')),
            'PERSONGROUP': clean_value(r.get('PERSONGROUP', '')),
            'BASE_DO_PERFIL': clean_value(norm_env(env)),
            'GRUPOS_ACESSO': '; '.join(grupos),
            'DATA_ULTIMO_ACESSO': last_login[0].strftime('%Y-%m-%d %H:%M:%S') if last_login else '',
            'BASE_LOGADA': clean_value(norm_env(last_login[1])) if last_login else '',
            'DIAS_SEM_ACESSO': dias_sem_acesso,
            'ALERTA_INATIVIDADE': alerta,
        })

    # Ativos primeiro, depois inativos; dentro dos ativos, quem tem alerta de
    # inatividade sobe para o topo (é o que a auditoria mais precisa ver
    # primeiro); resto ordenado por login mais antigo primeiro.
    rows.sort(key=lambda x: (
        x['STATUS'].upper() != 'ACTIVE',
        x['ALERTA_INATIVIDADE'] == '',
        x['DATA_ULTIMO_ACESSO'] == '',
        x['DATA_ULTIMO_ACESSO'],
    ), reverse=False)

    add_sheet('16_Auditoria_Acesso', headers, rows)
    print(f'✓ Aba 16 adicionada: {len(rows)} registros (ativos + inativados)')


# --- Allocation Excel Sheets (Maximo 9) ---
def add_allocation_sheets(wb, allocation_data, add_sheet):
    """Adiciona abas do Excel com o saneamento de alocação (Maximo 9)."""
    stats = allocation_data['stats']
    analises = allocation_data['analises']

    # Aba 21: Resumo de Alocação
    summary_rows = [
        ['Métrica', 'Valor'],
        ['Usuários analisados', stats['total_users']],
        ['Usuários com login nos últimos 90d', stats['users_with_logins_90d']],
        ['Usuários inativos (no Maximo)', stats['users_inactive']],
        ['Usuários com STATUS ausente na extração (revisar)', stats.get('users_status_desconhecido', 0)],
        ['Usuários que exigem conta em >1 ambiente', stats['users_multi_env']],
        ['Total de contas sugeridas (soma)', stats['total_suggested_accounts']],
        ['Limite mín. de acessos p/ ambiente secundário', stats['min_secundario']],
        ['Janela de análise (início)', stats['window_start']],
        ['Janela de análise (fim)', stats['window_end']],
    ]
    add_sheet('12_Alocacao_Resumo', ['Métrica', 'Valor'], summary_rows)

    # Aba 22: Detalhamento de Alocação / Sugestão (com colunas individuais por ambiente)
    # Precisa bater com allocation_analyzer.KNOWN_ENVS — um ambiente que está em
    # KNOWN_ENVS mas não aqui é excluído do cálculo de 'OUTROS' (que só soma
    # ambientes NÃO listados em KNOWN_ENVS) e também não ganha coluna própria,
    # desaparecendo silenciosamente da aba se algum dia tiver login registrado.
    ENV_COLS = ['BASE', 'ODN1', 'ODN2', 'ODN3', 'ODN4', 'N06', 'N08', 'N09', 'HTQ', 'POL', 'PGA', 'PGB', 'PGC', 'OUTROS']
    headers = (['USERID', 'NOME', 'STATUS', 'EMAIL', 'ALOCACAO_PRINCIPAL',
                'AMBIENTE_PRINCIPAL_USO', 'LOGINS_90D', 'ULTIMO_LOGIN'] +
               ENV_COLS +
               ['AMBIENTES_SECUNDARIOS', 'CONTAS_SUGERIDAS', 'HISTORICO_90D', 'MOTIVO'])
    rows = []
    for a in analises:
        detail = a.get('env_logins_detail', {})
        env_vals = [detail.get(e, 0) for e in ENV_COLS]
        rows.append({
            'USERID': a['userid'],
            'NOME': a['displayname'],
            'STATUS': a['status'],
            'EMAIL': a['email'],
            'ALOCACAO_PRINCIPAL': a['allocation_primary'],
            'AMBIENTE_PRINCIPAL_USO': a['primary_env'],
            'LOGINS_90D': a['total_logins_90d'],
            'ULTIMO_LOGIN': a['last_login'],
            **dict(zip(ENV_COLS, env_vals)),
            'AMBIENTES_SECUNDARIOS': '; '.join(a['secondary_envs']),
            'CONTAS_SUGERIDAS': '; '.join(a['suggested_accounts']),
            'HISTORICO_90D': a['detail'],
            'MOTIVO': a['reason'],
        })
    add_sheet('13_Alocacao_Sugestao', headers, rows)


# --- Migration Excel Sheets ---
def add_migration_sheets(wb, migration_data, add_sheet):
    """Adiciona a aba de recomendações de migração. O resumo por tipo foi
    consolidado na aba 17_Resumo_Consolidado (pedido do usuário 2026-07-11:
    menos abas, informação fundida)."""

    # Lista Completa de Recomendações — o entregável de migração MAS 9:
    # migrar?, qual grupo (atuais + recomendado por cargo), quais acessos.
    headers = ['Tipo', 'Prioridade', 'USERID', 'E-mail', 'Nome AD', 'Nome Maximo',
               'Status AD', 'Status Maximo', 'Ambientes (env:status)', 'Cargo',
               'Grupos Maximo Atuais', 'Grupo Recomendado (MAS 9)', 'Match Por',
               'Grupos AD', 'Motivo', 'Ação']
    rows = []
    for r in migration_data:
        rows.append({
            'Tipo': r['tipo'],
            'Prioridade': r['prioridade'],
            'USERID': r['userid'],
            'E-mail': r['email'],
            'Nome AD': r['nome_ad'],
            'Nome Maximo': r['nome_maximo'],
            'Status AD': r['status_ad'],
            'Status Maximo': r['status_maximo'],
            'Ambientes (env:status)': r.get('envs_detalhe') or r['envs'],
            'Cargo': r.get('cargo', ''),
            'Grupos Maximo Atuais': r.get('grupos_maximo', ''),
            'Grupo Recomendado (MAS 9)': r.get('grupo_recomendado_mas9', ''),
            'Match Por': r.get('match_por', ''),
            'Grupos AD': r['grupos_ad'],
            'Motivo': r['motivo'],
            'Ação': r['acao'],
        })
    add_sheet('11_Migracao_MAS9', headers, rows)


# --- Main Orchestration ---
def main():
    """
    Orchestrates the entire process of data loading, analysis, and report generation.
    """
    # 1. Load Data
    all_data = load_all_data()

    # --- INJEÇÃO DA NOVA ANÁLISE DE IDENTIDADE ---
    # Obtém os dados de identidade já processados e deduplicados
    identity_analytics = get_unique_users_data()

    # 2. Build User Profiles
    user_profiles = build_user_profiles(
        all_data["identities"],
        all_data["access_rows"],
        all_data["emails"],
        all_data["persons"],
        all_data["persongroupview"],
    )

    # 2b. Enrich with LOCATION_SITE from persongroupview (ENVIRONMENT column) - BEFORE simulation
    # Usa lógica inteligente: pega o ambiente do último login ou DEFSITE
    def _valid_site(v):
        """'0' é um placeholder de sitedefault/DEFSITE vazio no Maximo — como
        é uma string não-vazia, 'sitedefault or locationsite' o aceitava como
        válido e nunca caía pro próximo nível da cadeia de prioridade
        (auditoria 2026-07-14: 300+ usuários FORESEA ficavam com
        LOCATION_SITE='0' em vez do ambiente real)."""
        v = (v or '').strip()
        return v if v and v != '0' else ''

    persongroupview_map = {}
    for pgv in all_data.get("persongroupview", []):
        uid = str(pgv.get('personid', '')).strip().upper()
        env = pgv.get('ENVIRONMENT', '').strip()
        defsite = _valid_site(pgv.get('sitedefault', '')) or _valid_site(pgv.get('locationsite', ''))
        if uid and env:
            if uid not in persongroupview_map:
                persongroupview_map[uid] = {'environment': env, 'defsite': defsite}

    # Carregar logintracking para inferir ambiente real
    logintrack = load_csv(IN_DIR / 'consolidated_logintracking_from_sources.csv')

    # Inferir ambiente real do usuário baseado em CLIENTHOST
    def infer_env_from_clienthost(clienthost):
        if not clienthost:
            return None, False
        host = clienthost.strip().upper()
        if host.replace('.', '').isdigit():
            return None, True
        if 'ODRL-SP-SV' in host:
            return None, True
        if 'ODRL-ODN2-SV' in host:
            return 'ODN2', False
        if 'ODRL-ODN1-SV' in host:
            return 'ODN1', False
        if 'ODRL-N06-SV' in host:
            return 'N06', False
        if 'ODRL-N08-SV' in host:
            return 'N08', False
        if 'ODRL-N09-SV' in host:
            return 'N09', False
        if host.startswith('OD2-') or '-OD2-' in host:
            return 'ODN2', False
        if host.startswith('OD1-') or '-OD1-' in host:
            return 'ODN1', False
        if host.startswith('ON06-') or '-N06-' in host:
            return 'N06', False
        if host.startswith('ON08-') or '-N08-' in host:
            return 'N08', False
        if host.startswith('ON09-') or '-N09-' in host:
            return 'N09', False
        return None, False

    # Calcular ambiente real por usuário
    user_real_env = {}
    for rec in logintrack:
        if rec.get('ATTEMPTRESULT', '').strip().upper() != 'LOGIN':
            continue
        userid = rec.get('USERID', '').strip().upper()
        clienthost = rec.get('CLIENTHOST', '').strip()
        if userid:
            env, is_shared = infer_env_from_clienthost(clienthost)
            if env:
                user_real_env[userid] = env

    for profile in user_profiles.values():
        uid = str(profile.get('USERID', '')).strip().upper()
        # Cascata de prioridade — cada nível só "ganha" se produzir um valor
        # de verdade válido (não '0'); antes, um nível com chave presente mas
        # valor='0' vencia o elif e travava a cascata sem chegar nos
        # próximos níveis (auditoria 2026-07-14).
        pgv_entry = persongroupview_map.get(uid) or {}
        site = (
            _valid_site(user_real_env.get(uid, ''))  # Prioridade 1: ambiente real do logintracking
            or _valid_site(pgv_entry.get('defsite', ''))  # Prioridade 2a: DEFSITE do persongroupview
            or _valid_site(pgv_entry.get('environment', ''))  # Prioridade 2b: ambiente do persongroupview
            or _valid_site(profile.get('DEFSITE', ''))  # Prioridade 3: DEFSITE do próprio perfil
        )
        if site:
            profile['LOCATION_SITE'] = site

    active_profiles = [p for p in user_profiles.values() if p['STATUS'] == 'ACTIVE']

    # 3. Perform Governance Analysis
    domain_counts = analyze_governance(active_profiles)
    title_divergences_list, detailed_divergences = analyze_title_divergences(all_data["access_rows"], user_profiles)

    # 4. Perform AppPoints Simulation (por escopo)
    foresea_profiles = [
        p for p in active_profiles
        if p['DOMAIN_CATEGORY'] in ('FORESEA', 'PARCEIRO')
    ]
    other_profiles = [
        p for p in active_profiles
        if p['DOMAIN_CATEGORY'] not in ('FORESEA', 'PARCEIRO', 'SEM DOMINIO')
    ]
    missing_email_profiles = [
        p for p in active_profiles
        if p['DOMAIN_CATEGORY'] == 'SEM DOMINIO'
    ]

    missing_email_rows = [
        {
            'USERID': p['USERID'],
            'DISPLAYNAME': '; '.join(sorted(x for x in p['DISPLAYNAME'] if x)) or p['USERID'],
            'STATUS': p['STATUS'],
            'ENVS': '; '.join(sorted(p['ENVS'])),
            'TYPE': '; '.join(sorted(p['TYPE'])),
            'GROUPS_COUNT': len(p['GROUPS']),
            'GROUPS': '; '.join(sorted(p['GROUPS'])),
            'TITLES': '; '.join(sorted(p['TITLES'])),
            'PERSONGROUPS': '; '.join(sorted(p['PERSONGROUPS'])),
            'REVIEW_REASON': 'Sem email nominal/dominio valido no cadastro. Revisar antes de contar AppPoints.',
        }
        for p in missing_email_profiles
    ]

    foresea_app_points = simulate_app_points(foresea_profiles, user_real_env)
    other_app_points = simulate_app_points(other_profiles, user_real_env)

    # Re-enrich with LOCATION_SITE after simulation (simulate_app_points creates new dicts)
    for row in foresea_app_points + other_app_points:
        uid = str(row.get('USERID', '')).strip().upper()
        if not _valid_site(row.get('LOCATION_SITE', '')):
            pgv_entry = persongroupview_map.get(uid) or {}
            # Priorizar defsite (ambiente alocado) sobre environment (ambiente do registro)
            row['LOCATION_SITE'] = _valid_site(pgv_entry.get('defsite', '')) or _valid_site(pgv_entry.get('environment', ''))

    app_points_by_scope = {
        'foresea': foresea_app_points,
        'other': other_app_points,
    }

        # Gera o license decision plan com TODOS os usuários (incluindo SEM DOMINIO)
    # Usuários SEM DOMINIO são incluídos mas marcados para revisão
    sem_dominio_rows = [
        {
            **p,
            'DOMAIN_CATEGORY': 'SEM DOMINIO',
            'MIGRATION_SCOPE': 'REVIEW_MISSING_EMAIL',
            'LICENSE_MODEL': 'CONCURRENT',
            'ENTITLEMENT': 'BASE',
            'APP_POINTS': calculate_app_points('BASE', 'CONCURRENT'),
            'OPTIMIZATION_REC': 'REQUER_REVISAO',
            'OPTIMIZATION_REASON': 'Sem email cadastrado. Revisar antes de definir licença.'
        }
        for p in missing_email_profiles
    ]


    all_app_points_for_plan = foresea_app_points + other_app_points + sem_dominio_rows
    write_license_decision_plan(all_app_points_for_plan)

    # Compatibilidade com o pipeline/HTML atual (por enquanto ainda consumimos um único universo)
    # Depois este valor será substituído por dados por escopo no frontend.
    app_points_data_optimized = all_app_points_for_plan

    # 6. Prepare Data for HTML Builder
    summary_data = {
        'active_profiles_count': len(active_profiles),
        'title_divergence_count': len(title_divergences_list),
        'ceiling_limit': 1200,
        'app_points_summary': {
            'auth_users': [s for s in app_points_data_optimized if s['LICENSE_MODEL'] == 'AUTHORIZED'],
            'conc_users': [s for s in app_points_data_optimized if s['LICENSE_MODEL'] == 'CONCURRENT'],
            'premium_users': [s for s in app_points_data_optimized if s['ENTITLEMENT'] == 'PREMIUM'],
        },
    }


    governance_data = {
        'cross_env': all_data['cross_env'],
        'login_conflicts': all_data['login_conflicts'],
        'worklist': all_data['worklist'],
        'title_divergences_list': title_divergences_list, # List of just titles
        'detailed_divergences': detailed_divergences, # Detailed structure for section 5
        'identities': all_data['identities'], # Needed for some original metrics
        'access_rows': all_data['access_rows'], # Needed for some original metrics
        'user_profiles': user_profiles, # Pass consolidated profiles for detailed tables
        'persongroupview': all_data['persongroupview'], # For audit sheet
        'groupuser': all_data['groupuser'], # For audit sheet (grupos de acesso por USERID)
        'logintrack': logintrack, # For audit sheet (data/base do último acesso real)
        'maxuserstatus': all_data['maxuserstatus'], # For audit sheet (histórico real de status da conta)
    }

    # 7. Análise de Saneamento de Identidades (AD vs Maximo)
    print("\n" + "=" * 100)
    print("ANÁLISE DE SANEAMENTO DE IDENTIDADES (AD vs Maximo)")
    print("=" * 100)
    sanity_result = analyze_sanity()

    # 7b. Análise de Recomendações de Migração
    print("\n" + "=" * 100)
    print("ANÁLISE DE RECOMENDAÇÕES DE MIGRAÇÃO")
    print("=" * 100)
    migration_recommendations = analyze_migration()

    # 7c. Saneamento de Alocação (histórico de logins + sugestão de ambientes)
    print("\n" + "=" * 100)
    print("SANEAMENTO DE ALOCAÇÃO (Maximo 9)")
    print("=" * 100)
    allocation_result = analyze_allocation()

    # 7d. Auditoria de Segregação de Funções (Emissor x Aprovador em Compras)
    print("\n" + "=" * 100)
    print("AUDITORIA DE SEGREGAÇÃO DE FUNÇÕES (EMISSOR x APROVADOR)")
    print("=" * 100)
    security_audit_result = analyze_security_audit()

    # 7e. Perfil de Acesso por Cargo (excesso/falta de grupos vs. baseline do cargo)
    print("\n" + "=" * 100)
    print("PERFIL DE ACESSO POR CARGO (grupos padrão x desvios)")
    print("=" * 100)
    group_baseline_result = analyze_group_baseline()
    from scripts.domain.group_baseline import print_summary as print_group_baseline_summary
    print_group_baseline_summary(group_baseline_result)

    # 7f. Padronização de Acesso (cargo x grupo padrão único para todas as unidades)
    print("\n" + "=" * 100)
    print("PADRONIZAÇÃO DE ACESSO (cargo x grupo padrão — material para terceirizada)")
    print("=" * 100)
    role_standardization_result = analyze_role_standardization()
    from scripts.domain.role_standardization import print_summary as print_role_standardization_summary
    print_role_standardization_summary(role_standardization_result)

    # 7g. Cenário Conciliado de licenciamento (dimensionamento oficial MAS 9)
    print("\n" + "=" * 100)
    print("CENÁRIO CONCILIADO DE LICENCIAMENTO (AD ativo × Maximo ativo × uso real)")
    print("=" * 100)
    reconciliation_result = analyze_license_reconciliation()
    from scripts.domain.license_reconciliation import print_summary as print_reconciliation_summary
    print_reconciliation_summary(reconciliation_result)

    # 8. Build and Write HTML (com dados de AD, Maximo, Sanity, Migration, Allocation, Security Audit, Perfil de Cargo e Padronização)
    html_content = build_html_structure(
        summary_data,
        governance_data,
        app_points_data_optimized,
        domain_counts,
        identity_analytics,
        ad_users=all_data.get('ad_users', []),
        maximo_users=all_data.get('maximo_users', []),
        sanity_data=sanity_result,
        migration_data=migration_recommendations,
        allocation_data=allocation_result,
        security_audit_data=security_audit_result,
        group_baseline_data=group_baseline_result,
        role_standardization_data=role_standardization_result,
        reconciliation_data=reconciliation_result,
    )
    html_path = OUT_DIR / 'maximo_unified_dashboard.html'
    html_path.write_text(html_content, encoding='utf-8')
    print(f'WROTE {html_path.name}')
    write_excel_workbook(summary_data, governance_data, app_points_data_optimized, domain_counts, missing_email_rows, sanity_result, migration_recommendations, identities=all_data.get('identities', []), allocation_data=allocation_result, security_audit_data=security_audit_result, group_baseline_data=group_baseline_result, role_standardization_data=role_standardization_result, reconciliation_data=reconciliation_result)

if __name__ == '__main__':
    main()