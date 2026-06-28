from openpyxl import load_workbook

wb = load_workbook('output/reports/maximo_risk_and_optimization_workbook.xlsx', read_only=True)
ws = wb['2_LicenseDecisionPlan']
headers = [cell.value for cell in ws[1]]
print('HEADERS:', headers)

candidates = ['LOCATION','LOCATION_SITE','location_site','LOCATION SITE','LOCATION.SITE']
for cname in candidates:
    if cname in headers:
        col = headers.index(cname) + 1
        total = 0
        nonempty = 0
        for r in ws.iter_rows(min_row=2, min_col=col, max_col=col, values_only=True):
            total += 1
            if r[0] not in (None, '', 'UNKNOWN'):
                nonempty += 1
        print(f"{cname}: {nonempty}/{total} non-empty")

# Also print sample of first 10 rows for LOCATION and a potential LOCATION_SITE
print('\nSample rows:')
loc_col = headers.index('LOCATION')+1 if 'LOCATION' in headers else None
ls_col = None
for cand in ['LOCATION_SITE','location_site','LOCATION SITE']:
    if cand in headers:
        ls_col = headers.index(cand)+1
        break

for i, row in enumerate(ws.iter_rows(min_row=2, max_row=50, values_only=True), start=2):
    if i>12: break
    out = f"Row {i}:"
    if loc_col:
        out += f" LOCATION={row[loc_col-1]}"
    if ls_col:
        out += f" LOCATION_SITE={row[ls_col-1]}"
    print(out)
