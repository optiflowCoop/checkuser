#!/usr/bin/env python3
"""Verifica se ALESSANDROCORREA está corretamente atualizado no XLSX"""

from openpyxl import load_workbook
from pathlib import Path

REPORTS_DIR = Path("output/reports")
wb = load_workbook(REPORTS_DIR / 'maximo_risk_and_optimization_workbook.xlsx')
ws = wb['2_LicenseDecisionPlan']

# Find ALESSANDROCORREA
headers = [cell.value for cell in ws[1]]
userid_col = next((i for i, h in enumerate(headers, 1) if h == 'USERID'), None)
location_col = next((i for i, h in enumerate(headers, 1) if h == 'LOCATION'), None)
presence_col = next((i for i, h in enumerate(headers, 1) if h == 'OPERATIONAL_PRESENCE'), None)

print(f"Colunas: USERID={userid_col}, LOCATION={location_col}, OPERATIONAL_PRESENCE={presence_col}")
print(f"Total de colunas: {ws.max_column}")
print(f"Total de linhas: {ws.max_row}")

found = False
for row_idx in range(2, ws.max_row + 1):
    userid = ws.cell(row=row_idx, column=userid_col).value
    if userid and str(userid).upper() == 'ALESSANDROCORREA':
        location = ws.cell(row=row_idx, column=location_col).value if location_col else None
        presence = ws.cell(row=row_idx, column=presence_col).value if presence_col else None
        displayname = ws.cell(row=row_idx, column=2).value if 2 <= ws.max_column else None
        
        print(f"\n✅ ALESSANDROCORREA encontrado na linha {row_idx}:")
        print(f"  DISPLAYNAME: {displayname}")
        print(f"  LOCATION: {location}")
        print(f"  OPERATIONAL_PRESENCE: {presence}")
        found = True
        break

if not found:
    print("\n❌ ALESSANDROCORREA não encontrado no XLSX")
