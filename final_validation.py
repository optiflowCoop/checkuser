#!/usr/bin/env python3
"""
Script de validação final - Demonstra que as inconsistências foram resolvidas
"""
import csv
from pathlib import Path
from openpyxl import load_workbook

print("=" * 70)
print("VALIDAÇÃO FINAL - CORREÇÃO DE INCONSISTÊNCIAS")
print("=" * 70)

# 1. Verificar ALESSANDROCORREA em usage_analysis
print("\n1. VERIFICAÇÃO EM usage_analysis_phase3.csv")
print("-" * 70)

usage_path = Path("output/consolidated/usage_analysis_phase3.csv")
with open(usage_path, encoding='utf-8-sig', errors='replace') as f:
    reader = csv.DictReader(f)
    for row in reader:
        userid = str(row.get('USERID', '')).strip().upper()
        if userid == 'ALESSANDROCORREA':
            print(f"✓ ALESSANDROCORREA encontrado")
            print(f"  - LOCATION: {row.get('LOCATION')} (fonte: PersonGroupView)")
            print(f"  - OPERATIONAL_PRESENCE: {row.get('OPERATIONAL_PRESENCE')}")
            print(f"  - LOGIN_COUNT_90D: {row.get('LOGIN_COUNT_90D')}")
            print(f"  - DAYS_SINCE_LAST: {row.get('DAYS_SINCE_LAST')}")
            print(f"  - USER_TIER: {row.get('USER_TIER')}")
            break

# 2. Verificar no XLSX
print("\n2. VERIFICAÇÃO NO XLSX - Sheet '2_LicenseDecisionPlan'")
print("-" * 70)

wb = load_workbook(Path("output/reports/maximo_risk_and_optimization_workbook.xlsx"))
ws = wb['2_LicenseDecisionPlan']

headers = [cell.value for cell in ws[1]]
userid_col = next((i for i, h in enumerate(headers, 1) if h == 'USERID'), None)
location_col = next((i for i, h in enumerate(headers, 1) if h == 'LOCATION'), None)
presence_col = next((i for i, h in enumerate(headers, 1) if h == 'OPERATIONAL_PRESENCE'), None)

print(f"Colunas identificadas:")
print(f"  - USERID: coluna {userid_col}")
print(f"  - LOCATION: coluna {location_col}")
print(f"  - OPERATIONAL_PRESENCE: coluna {presence_col}")

for row_idx in range(2, ws.max_row + 1):
    userid = ws.cell(row=row_idx, column=userid_col).value
    if userid and str(userid).upper() == 'ALESSANDROCORREA':
        location = ws.cell(row=row_idx, column=location_col).value if location_col else None
        presence = ws.cell(row=row_idx, column=presence_col).value if presence_col else None
        displayname = ws.cell(row=row_idx, column=2).value
        
        print(f"\n✓ ALESSANDROCORREA encontrado na linha {row_idx}")
        print(f"  - DISPLAYNAME: {displayname}")
        print(f"  - LOCATION: {location} ✅")
        print(f"  - OPERATIONAL_PRESENCE: {presence} ✅")
        
        # Validações
        if location == 'N08' and presence == 'OFFSHORE':
            print(f"\n✅ VALIDAÇÃO PASSOU - Dados consistentes!")
        else:
            print(f"\n❌ VALIDAÇÃO FALHOU - Dados inconsistentes!")
        break

# 3. Amostra de outros usuários
print("\n3. AMOSTRA DE OUTROS USUÁRIOS COM LOCATION ATUALIZADA")
print("-" * 70)

sample_count = 0
for row_idx in range(2, min(ws.max_row + 1, 50)):
    userid = ws.cell(row=row_idx, column=userid_col).value
    location = ws.cell(row=row_idx, column=location_col).value if location_col else None
    presence = ws.cell(row=row_idx, column=presence_col).value if presence_col else None
    
    if location and location != 'UNKNOWN':
        print(f"  {str(userid):20} | LOCATION: {location:15} | PRESENCE: {presence}")
        sample_count += 1
        if sample_count >= 5:
            break

print("\n" + "=" * 70)
print("FIM DA VALIDAÇÃO")
print("=" * 70)
